import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import shutil

# Create folder for Excel matrices
excel_folder = r"C:\Users\Petugas\Claude\Projects\Pengadaan\EXCEL_MATRICES"
if not os.path.exists(excel_folder):
    os.makedirs(excel_folder)
    print(f"[OK] Folder created: {excel_folder}")

# ==================== MATRIX 1: RACI ====================
print("\n[CREATING] Matrix 1: RACI Responsibility...")

wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "RACI"

activities = [
    "Budget Planning", "Budget Approval", "Vendor Registration", "Vendor Qualification",
    "Create PP", "Approve PP (<10M)", "Approve PP (10-25M)", "Create SPPH", "Evaluate Bids",
    "Create PO", "Approve PO", "Send PO to Vendor", "Create SPK/PKS", "Goods Delivery",
    "Create BAPB", "QC Inspection", "Approve BAPB", "Invoice Submission", "3-Way Match",
    "Approve Payment", "Vendor Scoring", "Monthly Review", "Compliance Audit"
]

stakeholders = ["Dir Utama", "Dir Ops", "Dir Keuangan", "Manager", "Kasie Barang", "Kasie Jasa",
                "GM", "Finance Mgr", "SBU Head", "Lab Mgr", "Pharmacy", "Warehouse", "QC", "SPI", "Vendor"]

raci_data = {
    "Budget Planning": ["C", "C", "A", "I", "I", "I", "C", "C", "C", "I", "I", "I", "I", "C", "I"],
    "Budget Approval": ["A", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Vendor Registration": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "C", "R"],
    "Vendor Qualification": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "C", "I"],
    "Create PP": ["I", "I", "I", "I", "I", "I", "I", "I", "R", "C", "C", "I", "I", "I", "I"],
    "Approve PP (<10M)": ["I", "I", "I", "R", "R", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Approve PP (10-25M)": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Create SPPH": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Evaluate Bids": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Create PO": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Approve PO": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Send PO to Vendor": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "R"],
    "Create SPK/PKS": ["I", "I", "I", "I", "C", "R", "C", "I", "I", "I", "I", "I", "I", "I", "R"],
    "Goods Delivery": ["I", "I", "I", "I", "I", "I", "I", "I", "I", "R", "R", "R", "I", "I", "R"],
    "Create BAPB": ["I", "I", "I", "I", "I", "I", "I", "I", "C", "R", "R", "R", "R", "I", "I"],
    "QC Inspection": ["I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "C", "R", "I", "I"],
    "Approve BAPB": ["I", "I", "I", "I", "I", "I", "I", "I", "C", "R", "R", "C", "R", "I", "I"],
    "Invoice Submission": ["I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "R"],
    "3-Way Match": ["I", "I", "I", "C", "I", "I", "I", "R", "I", "I", "I", "I", "I", "I", "I"],
    "Approve Payment": ["I", "I", "R", "C", "I", "I", "I", "R", "I", "I", "I", "I", "I", "I", "I"],
    "Vendor Scoring": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "C", "I"],
    "Monthly Review": ["I", "I", "I", "R", "C", "C", "C", "R", "I", "I", "I", "I", "I", "C", "I"],
    "Compliance Audit": ["I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "R", "I"],
}

# Setup colors & styles
color_r = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
color_a = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
color_c = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
color_i = PatternFill(start_color="4D96FF", end_color="4D96FF", fill_type="solid")
color_header = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")
font_white = Font(bold=True, color="FFFFFF", size=11)
font_black = Font(bold=True, color="000000", size=10)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Title
ws1.merge_cells('A1:P1')
title_cell = ws1['A1']
title_cell.value = "RACI MATRIX - PENGADAAN KMU"
title_cell.font = font_white
title_cell.fill = color_header
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 25

# Legend
ws1.merge_cells('A2:P2')
legend_cell = ws1['A2']
legend_cell.value = "R = Responsible (Do) | A = Accountable (Approve) | C = Consulted (Ask) | I = Informed (Tell)"
legend_cell.font = Font(italic=True, size=9)
ws1.row_dimensions[2].height = 20

# Headers
ws1['A4'] = "ACTIVITY"
ws1['A4'].font = font_white
ws1['A4'].fill = color_header
ws1['A4'].border = border

for col, stakeholder in enumerate(stakeholders, start=2):
    cell = ws1.cell(row=4, column=col)
    cell.value = stakeholder
    cell.font = font_white
    cell.fill = color_header
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Data rows
for row_idx, activity in enumerate(activities, start=5):
    cell = ws1.cell(row=row_idx, column=1)
    cell.value = activity
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    cell.border = border

    raci_row = raci_data.get(activity, [])
    for col_idx, value in enumerate(raci_row, start=2):
        cell = ws1.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = font_black
        cell.border = border

        if value == "R":
            cell.fill = color_r
        elif value == "A":
            cell.fill = color_a
        elif value == "C":
            cell.fill = color_c
        elif value == "I":
            cell.fill = color_i

# Column widths
ws1.column_dimensions['A'].width = 25
for col_idx in range(2, len(stakeholders) + 2):
    ws1.column_dimensions[get_column_letter(col_idx)].width = 12

for row in ws1.iter_rows():
    ws1.row_dimensions[row[0].row].height = 20

ws1.freeze_panes = 'B5'
wb1.save(os.path.join(excel_folder, "MATRIX_1_RACI_RESPONSIBILITY.xlsx"))
print("[OK] Matrix 1 saved")

# ==================== MATRIX 2: AUTHORITY APPROVAL ====================
print("[CREATING] Matrix 2: Authority Approval...")

wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Authority"

ws2['A1'] = "APPROVAL AUTHORITY MATRIX - PENGADAAN KMU"
ws2['A1'].font = font_white
ws2['A1'].fill = color_header
ws2.merge_cells('A1:F1')
ws2.row_dimensions[1].height = 25

headers2 = ["Amount Range", "Daan Umum", "Daan Jasa", "Required Approvers", "SLA", "Notes"]
for col, header in enumerate(headers2, start=1):
    cell = ws2.cell(row=3, column=col)
    cell.value = header
    cell.font = font_white
    cell.fill = color_header
    cell.border = border
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

data2 = [
    ["≤ Rp 10 Juta", "Kasie + Manager", "Kasie + Manager", "Kasie + Manager", "2 jam", "Fast-track"],
    ["Rp 10-25 Juta", "Kasie + Manager + GM", "Kasie + Manager + GM", "3 levels", "4 jam", "Standard"],
    ["Rp 25-50 Juta", "+ Dir Ops", "+ Dir Ops", "4 levels", "8 jam", "Multi-level"],
    ["Rp 50-100 Juta", "+ Dir Utama", "+ Dir Utama", "5 levels", "12 jam", "Highest"],
    ["> Rp 100 Juta", "Needs board review", "Dir Utama + Bank guarantee", "Special", "24 jam", "Board approval"],
]

for row_idx, row_data in enumerate(data2, start=4):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='left', wrap_text=True)

ws2.column_dimensions['A'].width = 20
for col in ['B', 'C', 'D', 'E', 'F']:
    ws2.column_dimensions[col].width = 18

wb2.save(os.path.join(excel_folder, "MATRIX_2_AUTHORITY_APPROVAL.xlsx"))
print("[OK] Matrix 2 saved")

# ==================== MATRIX 3: VENDOR SCORING ====================
print("[CREATING] Matrix 3: Vendor Scoring Criteria...")

wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = "Vendor Scoring"

ws3['A1'] = "VENDOR SCORING CRITERIA MATRIX"
ws3['A1'].font = font_white
ws3['A1'].fill = color_header
ws3.merge_cells('A1:C1')

headers3 = ["Dimension", "Weight %", "Sub-Criteria (1-5 Scale)"]
for col, header in enumerate(headers3, start=1):
    cell = ws3.cell(row=3, column=col)
    cell.value = header
    cell.font = font_white
    cell.fill = color_header
    cell.border = border

data3 = [
    ["HARGA (Price)", "25%", "Price competitiveness, volume discounts"],
    ["KUALITAS (Quality)", "20%", "Product quality, certifications, complaint rate"],
    ["PENGIRIMAN (Delivery)", "20%", "On-time rate, consistency, flexibility"],
    ["KEPATUHAN (Compliance)", "15%", "SPO adherence, documentation, SPA/Insurance"],
    ["KEMITRAAN (Partnership)", "10%", "Communication, cooperation, value-add"],
    ["KEUANGAN (Financial)", "10%", "Stability, payment history, capacity"],
]

for row_idx, row_data in enumerate(data3, start=4):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws3.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='left', wrap_text=True)

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 50

wb3.save(os.path.join(excel_folder, "MATRIX_3_VENDOR_SCORING.xlsx"))
print("[OK] Matrix 3 saved")

# ==================== MATRIX 4: MODULE STAKEHOLDER ====================
print("[CREATING] Matrix 4: Module x Stakeholder...")

wb4 = openpyxl.Workbook()
ws4 = wb4.active
ws4.title = "Module-Stakeholder"

ws4['A1'] = "MODULE x STAKEHOLDER MATRIX"
ws4['A1'].font = font_white
ws4['A1'].fill = color_header
ws4.merge_cells('A1:F1')

modules = ["M1: Procurement Platform", "M2: Vendor Management", "M3: Vendor Portal",
           "M4: AI Guardian", "M5: Execution", "M6: Quality", "M7: Invoice", "M8: Payment",
           "M9: Reporting", "M10: Vendor DB", "M11: Dashboard", "M12: Evaluation"]

stakeholders_short = ["Direksi", "Dept. Pengadaan", "SBU", "Finance", "Vendor", "SPI"]

ws4['A3'] = "Module / Stakeholder"
for col, stakeholder in enumerate(stakeholders_short, start=2):
    ws4.cell(row=3, column=col).value = stakeholder
    ws4.cell(row=3, column=col).fill = color_header
    ws4.cell(row=3, column=col).font = font_white
    ws4.cell(row=3, column=col).border = border

for row_idx, module in enumerate(modules, start=4):
    ws4.cell(row=row_idx, column=1).value = module
    ws4.cell(row=row_idx, column=1).fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    ws4.cell(row=row_idx, column=1).border = border

    for col in range(2, len(stakeholders_short) + 2):
        ws4.cell(row=row_idx, column=col).value = "Primary"
        ws4.cell(row=row_idx, column=col).border = border
        ws4.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')

ws4.column_dimensions['A'].width = 25
for col in range(2, len(stakeholders_short) + 2):
    ws4.column_dimensions[get_column_letter(col)].width = 15

wb4.save(os.path.join(excel_folder, "MATRIX_4_MODULE_STAKEHOLDER.xlsx"))
print("[OK] Matrix 4 saved")

# ==================== MATRIX 5: PROCESS RISK ====================
print("[CREATING] Matrix 5: Process x Risk...")

wb5 = openpyxl.Workbook()
ws5 = wb5.active
ws5.title = "Process-Risk"

ws5['A1'] = "PROCESS x RISK MATRIX"
ws5['A1'].font = font_white
ws5['A1'].fill = color_header
ws5.merge_cells('A1:G1')

processes = ["Budget Planning", "Vendor Registration", "Vendor Qualification", "PP Creation",
             "Approval Workflow", "Tender Process", "Vendor Selection", "PO Generation",
             "Goods Delivery", "Quality Inspection", "Invoice Processing", "Payment Processing"]

risks = ["Compliance Risk", "Financial Risk", "Operational Risk", "Vendor Risk", "Security Risk", "Reputational Risk"]

ws5['A3'] = "Process / Risk Type"
for col, risk in enumerate(risks, start=2):
    ws5.cell(row=3, column=col).value = risk
    ws5.cell(row=3, column=col).fill = color_header
    ws5.cell(row=3, column=col).font = font_white
    ws5.cell(row=3, column=col).border = border
    ws5.cell(row=3, column=col).alignment = Alignment(horizontal='center', wrap_text=True)

for row_idx, process in enumerate(processes, start=4):
    ws5.cell(row=row_idx, column=1).value = process
    ws5.cell(row=row_idx, column=1).fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    ws5.cell(row=row_idx, column=1).border = border

    for col in range(2, len(risks) + 2):
        ws5.cell(row=row_idx, column=col).value = "Medium"
        ws5.cell(row=row_idx, column=col).border = border
        ws5.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')

ws5.column_dimensions['A'].width = 25
for col in range(2, len(risks) + 2):
    ws5.column_dimensions[get_column_letter(col)].width = 15

wb5.save(os.path.join(excel_folder, "MATRIX_5_PROCESS_RISK.xlsx"))
print("[OK] Matrix 5 saved")

# ==================== MATRIX 6: DATA FLOW ====================
print("[CREATING] Matrix 6: Data Flow...")

wb6 = openpyxl.Workbook()
ws6 = wb6.active
ws6.title = "Data Flow"

ws6['A1'] = "DATA FLOW & INFORMATION SYSTEM MATRIX"
ws6['A1'].font = font_white
ws6['A1'].fill = color_header
ws6.merge_cells('A1:F1')

headers6 = ["Data Point", "Source System", "Target System", "Frequency", "Volume", "Owner"]
for col, header in enumerate(headers6, start=1):
    ws6.cell(row=3, column=col).value = header
    ws6.cell(row=3, column=col).fill = color_header
    ws6.cell(row=3, column=col).font = font_white
    ws6.cell(row=3, column=col).border = border

data6 = [
    ["Budget Allocation", "Finance", "Procurement", "Quarterly", "12 SBUs", "Finance Mgr"],
    ["Vendor ID", "Vendor Portal", "Procurement", "One-time", "200+", "Manager"],
    ["PP Request", "Procurement", "Finance", "Per request", "100s/month", "Kasie"],
    ["PO Number", "Procurement", "All systems", "Per PO", "100s/month", "Finance"],
    ["BAPB", "Warehouse", "All systems", "Per delivery", "100s/month", "Warehouse"],
    ["Invoice", "Vendor", "Finance", "Per invoice", "100s/month", "Finance"],
    ["Payment", "Finance", "Bank", "Per payment", "100s/month", "Finance"],
    ["Vendor Score", "QC/Finance", "Procurement", "Monthly", "200+ vendors", "Manager"],
    ["Compliance", "All systems", "SPI", "Continuous", "10K+/month", "SPI"],
]

for row_idx, row_data in enumerate(data6, start=4):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws6.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='left', wrap_text=True)

for col_idx in range(1, 7):
    ws6.column_dimensions[get_column_letter(col_idx)].width = 20

wb6.save(os.path.join(excel_folder, "MATRIX_6_DATA_FLOW.xlsx"))
print("[OK] Matrix 6 saved")

# ==================== CLEANUP ====================
print("\n[CLEANUP] Removing old CSV files...")
csv_files = [
    "MATRIX_1_RACI_RESPONSIBILITY.csv",
    "MATRIX_2_AUTHORITY_APPROVAL.csv",
    "MATRIX_3_VENDOR_SCORING_CRITERIA.csv",
    "MATRIX_4_MODULE_STAKEHOLDER.csv",
    "MATRIX_5_PROCESS_RISK.csv",
    "MATRIX_6_DATA_FLOW_INFORMATION_SYSTEM.csv",
]

for csv_file in csv_files:
    csv_path = os.path.join(r"C:\Users\Petugas\Claude\Projects\Pengadaan", csv_file)
    if os.path.exists(csv_path):
        os.remove(csv_path)
        print(f"[DELETED] {csv_file}")

print("\n[SUMMARY]")
print(f"[OK] All 6 Excel matrices created in folder: EXCEL_MATRICES")
print(f"[OK] All CSV files deleted")
print(f"[READY] Open EXCEL_MATRICES folder to use matrices!")
