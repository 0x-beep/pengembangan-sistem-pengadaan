import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "RACI Matrix"

# Data - Activities and Stakeholders
activities = [
    "Budget Planning",
    "Budget Approval",
    "Vendor Registration",
    "Vendor Qualification",
    "Create PP",
    "Approve PP (<10M)",
    "Approve PP (10-25M)",
    "Create SPPH",
    "Evaluate Bids",
    "Create PO",
    "Approve PO",
    "Send PO to Vendor",
    "Create SPK/PKS",
    "Goods Delivery",
    "Create BAPB",
    "QC Inspection",
    "Approve BAPB",
    "Invoice Submission",
    "3-Way Match",
    "Approve Payment",
    "Vendor Scoring",
    "Monthly Review",
    "Compliance Audit",
]

stakeholders = [
    "Dir Utama",
    "Dir Ops",
    "Dir Keuangan",
    "Manager",
    "Kasie Barang",
    "Kasie Jasa",
    "GM",
    "Finance Mgr",
    "SBU Head",
    "Lab Mgr",
    "Pharmacy",
    "Warehouse",
    "QC",
    "SPI",
    "Vendor"
]

# RACI data (simplified version)
raci_data = {
    "Budget Planning": ["C", "C", "A", "I", "I", "I", "C", "C", "C", "I", "I", "I", "I", "C", "I"],
    "Budget Approval": ["A", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Vendor Registration": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "C", "R"],
    "Vendor Qualification": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "C", "I"],
    "Create PP": ["I", "I", "I", "I", "I", "I", "I", "I", "R", "C", "C", "I", "I", "I", "I"],
    "Approve PP (<10M)": ["I", "I", "I", "R", "R", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Approve PP (10-25M)": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Create SPPH": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Evaluate Bids": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Create PO": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Approve PO": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "I"],
    "Send PO to Vendor": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "I", "R"],
    "Create SPK/PKS": ["I", "I", "I", "I", "C", "R", "C", "I", "I", "I", "I", "I", "I", "I", "R"],
    "Goods Delivery": ["I", "I", "I", "I", "I", "I", "I", "I", "I", "R", "R", "R", "I", "I", "R"],
    "Create BAPB": ["I", "I", "I", "I", "I", "I", "I", "I", "C", "R", "R", "R", "R", "I", "I"],
    "QC Inspection": ["I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "C", "R", "I", "I"],
    "Approve BAPB": ["I", "I", "I", "I", "I", "I", "I", "I", "C", "R", "R", "C", "R", "I", "I"],
    "Invoice Submission": ["I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "R"],
    "3-Way Match": ["I", "I", "I", "C", "I", "I", "I", "R", "I", "I", "I", "I", "I", "I", "I"],
    "Approve Payment": ["I", "I", "R", "C", "I", "I", "I", "R", "I", "I", "I", "I", "I", "I", "I"],
    "Vendor Scoring": ["I", "I", "I", "R", "R", "R", "C", "I", "I", "I", "I", "I", "I", "C", "I"],
    "Monthly Review": ["I", "I", "I", "R", "C", "C", "C", "R", "I", "I", "I", "I", "I", "C", "I"],
    "Compliance Audit": ["I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "I", "R", "I"],
}

# Define colors
color_r = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red
color_a = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")  # Yellow
color_c = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")  # Green
color_i = PatternFill(start_color="4D96FF", end_color="4D96FF", fill_type="solid")  # Blue
color_header = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")  # Dark

font_white = Font(bold=True, color="FFFFFF", size=11)
font_black = Font(bold=True, color="000000", size=10)
font_header = Font(bold=True, color="FFFFFF", size=12)

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add title
ws.merge_cells('A1:P1')
title_cell = ws['A1']
title_cell.value = "RACI MATRIX - PENGADAAN KMU"
title_cell.font = font_header
title_cell.fill = color_header
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 25

# Add legend
ws.merge_cells('A2:P2')
legend_cell = ws['A2']
legend_cell.value = "R = Responsible (Do) | A = Accountable (Approve) | C = Consulted (Ask) | I = Informed (Tell)"
legend_cell.font = Font(italic=True, size=9)
ws.row_dimensions[2].height = 20

# Add headers
row = 4
ws['A4'] = "ACTIVITY"
ws['A4'].font = font_white
ws['A4'].fill = color_header
ws['A4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col, stakeholder in enumerate(stakeholders, start=2):
    cell = ws.cell(row=row, column=col)
    cell.value = stakeholder
    cell.font = font_white
    cell.fill = color_header
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Add data
for row_idx, activity in enumerate(activities, start=5):
    # Activity name
    cell = ws.cell(row=row_idx, column=1)
    cell.value = activity
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = border

    # RACI values
    raci_row = raci_data.get(activity, [])
    for col_idx, value in enumerate(raci_row, start=2):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = font_black
        cell.border = border

        # Color coding
        if value == "R":
            cell.fill = color_r
        elif value == "A":
            cell.fill = color_a
        elif value == "C":
            cell.fill = color_c
        elif value == "I":
            cell.fill = color_i

# Adjust column widths
ws.column_dimensions['A'].width = 25
for col_idx in range(2, len(stakeholders) + 2):
    ws.column_dimensions[get_column_letter(col_idx)].width = 12

# Set row height
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 20

# Freeze panes (freeze first row and first column)
ws.freeze_panes = 'B5'

# Save
output_path = r"C:\Users\Petugas\Claude\Projects\Pengadaan\MATRIX_1_RACI_RESPONSIBILITY.xlsx"
wb.save(output_path)
print(f"[OK] Excel file created: {output_path}")
print(f"[FORMAT] Proper columns, color-coded (R=Red, A=Yellow, C=Green, I=Blue)")
print(f"[FROZEN] Header row and Activity column")
